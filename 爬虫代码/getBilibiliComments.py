import requests
import math
import json
from openpyxl import Workbook
# url="https://api.bilibili.com/x/v2/reply?jsonp=jsonp&type=1&oid=458918620&pn="会缺部分数据
url='https://api.bilibili.com/x/v2/reply?jsonp=jsonp&type=1&oid=412980008&sort=2&pn='
#获取json数据
def download(url):
    result=requests.get(url)
    return json.loads(result.text)["data"]
#excel数据导出
def excel_out(data):
    global  index
    ws.cell(row=index,column=1,value=index-1)
    ws.cell(row=index,column=2,value=data["member"]["uname"])
    ws.cell(row=index,column=3,value=data["member"]["level_info"]["current_level"])
    ws.cell(row=index,column=4,value=data["member"]["sex"])
    ws.cell(row=index,column=5,value=data["content"]["message"])
    ws.cell(row=index,column=6,value=data["like"])
    index+=1
index=2;
wb=Workbook()
ws=wb.active
ws.title="共享经济-老蒋"
ws.cell(row=1,column=1,value="序号")
ws.cell(row=1,column=2,value="用户名")
ws.cell(row=1,column=3,value="等级")
ws.cell(row=1,column=4,value="性别")
ws.cell(row=1,column=5,value="评论")
ws.cell(row=1,column=6,value="点赞数")
data=download(url+"1")
totalPage=math.ceil(data["page"]["count"]/data["page"]["size"])
print("total: %s页" % totalPage)
for i in range(1,totalPage+1):
    data=download(url+str(i))
    replies=data["replies"]
    for v in replies:
        excel_out(v)
    print("当前进度 %.2f%%" % (i/totalPage*100))
wb.save("C:/Users/-zwx-/Desktop/Courses/Data Science/D/评论-【互联网简史06】共享经济之后再无风口（下）.xlsx")
print("完成\（*.*）/")
